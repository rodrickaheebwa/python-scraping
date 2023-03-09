from openpyxl import Workbook

my_wb = Workbook()
print('Sheet names: ', my_wb.sheetnames)

# creating new sheets
my_wb.create_sheet()
my_wb.create_sheet('my sheet', 1)
print('Sheet names: ', my_wb.sheetnames)

# renaming sheets
first_sheet = my_wb.active
first_sheet.title = 'first sheet'
print(first_sheet)

# activate sheet
my_sheet = my_wb['my sheet']
print(my_sheet)

# working with a sheet
first_sheet['A1'] = 2
first_sheet['C2'] = 4
print(first_sheet['A1'].value)
for value in first_sheet.values:
    print(value)

my_sheet['A1'] = 10
my_sheet['C2'] = 55
print(my_sheet['A1'].value)
for value in my_sheet.values:
    print(value)

# add a row of data
# sheet.append('col_1_value', 'col_2_value', 'col_3_value')
my_sheet.append([7,8,9])
first_sheet.append([3,6,9])
for value in first_sheet.values:
    print(value)

for value in my_sheet.values:
    print(value)

# saving to file to create it (otherwise it stays in memory)
# my_wb.save('.\\test\\excel_workbook.xlsx')